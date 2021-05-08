import openpyxl  # 导入包
from openpyxl import Workbook
import ship_online_rate

# 打开excle
file_name = r'D:/船舶在线率对比.xlsx'
wb = openpyxl.load_workbook(file_name)

# 选择要操作的sheet页
ws = wb['Sheet1']

for i in range(2, 2853):
    status = ws.cell(i, 6).value
    if status != '升级成功':
        continue
    try:
        ship_name = ws.cell(i, 3).value
        ship = ship_online_rate.ShipOnline()
        mmsi = ship.get_ship_mmsi(ship_name)
        if mmsi is None:
            ws.cell(i, 8).value = '船名和mmsi不匹配!'
            wb.save(file_name)
            continue

        data = ship.ship_online_rate(mmsi, '2021-04-23 00:00:00', '2021-04-30 23:59:59')
        online = data['online']
        ws.cell(i, 8).value = online

        data1 = ship.ship_online_rate(mmsi, '2021-05-07 00:00:00', '2021-05-08 23:59:59')
        online1 = data1['online']
        ws.cell(i, 9).value = online1
        wb.save(file_name)
    except Exception:
        ws.cell(i, 8).value = '没查询到对应的在线率!'
        wb.save(file_name)
wb.save(file_name)
